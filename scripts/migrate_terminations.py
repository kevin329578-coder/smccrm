import sys, os, json, re, datetime, warnings
sys.stdout.reconfigure(encoding='utf-8')
warnings.filterwarnings('ignore')
import openpyxl
import urllib.request
import urllib.parse

XLSX_PATH = r"C:\Users\SMC\Desktop\케빈의 프로젝트\SMC_CRM\캡시 사업운영 관리.xlsx"
ENV_PATH = os.path.join(os.path.dirname(__file__), '..', '.env.local')

env = {}
with open(ENV_PATH, encoding='utf-8') as f:
    for line in f:
        if '=' in line:
            k, v = line.strip().split('=', 1)
            env[k] = v

SUPABASE_URL = env['NEXT_PUBLIC_SUPABASE_URL']
SERVICE_KEY = env['SUPABASE_SERVICE_ROLE_KEY']
HEADERS = {
    'apikey': SERVICE_KEY, 'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json', 'Prefer': 'return=representation',
}
DATE_RE = re.compile(r'(\d{4})[.\-](\d{2})[.\-](\d{2})')

def to_text(v):
    return str(v).strip() if v is not None and str(v).strip() else None

def to_date_flexible(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()[:10]
    if isinstance(v, str):
        m = DATE_RE.match(v.strip())
        if m:
            return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None

def patch_json(url, payload):
    req = urllib.request.Request(url, data=json.dumps(payload).encode('utf-8'), method='PATCH', headers=HEADERS)
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())

def post_json(url, payload):
    req = urllib.request.Request(url, data=json.dumps(payload).encode('utf-8'), method='POST', headers=HEADERS)
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())

def fetch_all(url_base):
    results, offset = [], 0
    while True:
        req = urllib.request.Request(url_base + f'&limit=1000&offset={offset}', headers=HEADERS)
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read())
        if not data:
            break
        results.extend(data)
        offset += 1000
    return results

def build_alias_map(wb, franchisee_sheet_name):
    ws = wb[franchisee_sheet_name]
    m = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None or not row[7]:
            continue
        m[str(row[3]).strip()] = str(row[7]).strip()
    return m

def migrate_vehicle_terminations(sheet_name, franchisee_sheet_name, region):
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    alias_map = build_alias_map(wb, franchisee_sheet_name)
    r = urllib.parse.quote(region)
    franchisees = fetch_all(f"{SUPABASE_URL}/rest/v1/franchisees?region=eq.{r}&type=eq.%EB%B2%95%EC%9D%B8&select=id,business_name")
    fid_map = {f['business_name']: f['id'] for f in franchisees}

    ws = wb[sheet_name]
    updated, created, skipped = 0, 0, 0
    log_path = os.path.join(os.path.dirname(__file__), 'migrate_log.txt')
    with open(log_path, 'a', encoding='utf-8') as log:
        log.write(f"\n=== {sheet_name} ===\n")
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] is None or not row[3]:
                skipped += 1
                continue
            plate_no = to_text(row[3])
            terminated_at = to_date_flexible(row[5])
            payload = {'status': '해지'}
            if terminated_at:
                payload['terminated_at'] = terminated_at
            try:
                encoded_plate = urllib.parse.quote(plate_no)
                url = f"{SUPABASE_URL}/rest/v1/vehicles?plate_no=eq.{encoded_plate}"
                result = patch_json(url, payload)
                if result:
                    updated += 1
                    continue
                # 매칭되는 차량이 없으면 새로 생성 (해지 이력 시트가 현재 스냅샷보다 더 오래된 이력 포함)
                alias = to_text(row[2])
                business_name = alias_map.get(alias) or alias  # 별칭 매핑에 없으면 원래 시트값을 사업자명으로 간주(누락 가맹점 보완용)
                franchisee_id = fid_map.get(business_name) if business_name else None
                if not franchisee_id:
                    skipped += 1
                    log.write(f"가맹점 매칭 실패(신규 차량): {alias} (차량 {plate_no})\n")
                    continue
                create_payload = {
                    'franchisee_id': franchisee_id, 'plate_no': plate_no,
                    'status': '해지', 'terminated_at': terminated_at, 'notes': to_text(row[6]),
                }
                post_json(f"{SUPABASE_URL}/rest/v1/vehicles", create_payload)
                created += 1
            except Exception as e:
                skipped += 1
                log.write(f"실패: {plate_no} - {e}\n")
    print(f'{sheet_name} 완료. 업데이트 {updated}건, 신규생성 {created}건, 스킵 {skipped}건.')

def migrate_driver_terminations(sheet_name, franchisee_sheet_name, region):
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    alias_map = build_alias_map(wb, franchisee_sheet_name)
    r = urllib.parse.quote(region)
    franchisees = fetch_all(f"{SUPABASE_URL}/rest/v1/franchisees?region=eq.{r}&type=eq.%EB%B2%95%EC%9D%B8&select=id,business_name")
    fid_map = {f['business_name']: f['id'] for f in franchisees}

    ws = wb[sheet_name]
    updated, created, skipped = 0, 0, 0
    log_path = os.path.join(os.path.dirname(__file__), 'migrate_log.txt')
    with open(log_path, 'a', encoding='utf-8') as log:
        log.write(f"\n=== {sheet_name} ===\n")
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] is None or not row[4]:
                skipped += 1
                continue
            plate_no = to_text(row[4])
            name = to_text(row[3])
            reason = to_text(row[7]) if len(row) > 7 else None
            terminated_at = to_date_flexible(row[6])
            payload = {'call_status': '해지'}
            if terminated_at:
                payload['call_terminated_at'] = terminated_at
            if reason:
                payload['call_termination_reason'] = reason
            try:
                encoded_plate = urllib.parse.quote(plate_no)
                vlist = fetch_all(f"{SUPABASE_URL}/rest/v1/vehicles?plate_no=eq.{encoded_plate}&select=id,franchisee_id")
                vehicle_id = vlist[0]['id'] if vlist else None
                franchisee_id = vlist[0]['franchisee_id'] if vlist else None

                if vehicle_id:
                    url = f"{SUPABASE_URL}/rest/v1/drivers?vehicle_id=eq.{vehicle_id}"
                    result = patch_json(url, payload)
                    if result:
                        updated += 1
                        continue

                # 기사(또는 차량 자체)가 없으면 새로 생성
                if not franchisee_id:
                    alias = to_text(row[2])
                    business_name = alias_map.get(alias) or alias
                    franchisee_id = fid_map.get(business_name) if business_name else None
                if not franchisee_id:
                    skipped += 1
                    log.write(f"가맹점 매칭 실패(신규 기사): {row[2]} (기사 {name}, 차량 {plate_no})\n")
                    continue
                if not vehicle_id and plate_no:
                    v_created = post_json(f"{SUPABASE_URL}/rest/v1/vehicles", {
                        'franchisee_id': franchisee_id, 'plate_no': plate_no, 'status': '해지',
                    })
                    vehicle_id = v_created[0]['id']
                create_payload = {
                    'franchisee_id': franchisee_id, 'vehicle_id': vehicle_id, 'name': name,
                    'call_status': '해지', 'call_terminated_at': terminated_at,
                    'call_termination_reason': reason,
                }
                post_json(f"{SUPABASE_URL}/rest/v1/drivers", create_payload)
                created += 1
            except Exception as e:
                skipped += 1
                log.write(f"실패: {plate_no} - {e}\n")
    print(f'{sheet_name} 완료. 업데이트 {updated}건, 신규생성 {created}건, 스킵 {skipped}건.')

if __name__ == '__main__':
    migrate_vehicle_terminations('해지차량내역(서울법인)', '가맹점 정보(서울법인)', '서울')
    migrate_vehicle_terminations('해지차량내역(경기법인)', '가맹점 정보(경기법인)', '경기')
    migrate_driver_terminations('가맹콜해지내역(서울법인)', '가맹점 정보(서울법인)', '서울')
    migrate_driver_terminations('가맹콜해지내역(경기법인)', '가맹점 정보(경기법인)', '경기')
