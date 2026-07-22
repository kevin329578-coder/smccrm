import sys, os, json, re, datetime
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import urllib.request
import urllib.parse

XLSX_PATH = r"C:\Users\SMC\Desktop\케빈의 프로젝트\SMC_CRM\캡시 사업운영 관리.xlsx"
ENV_PATH = os.path.join(os.path.dirname(__file__), '..', '.env.local')

env = {}
with open(ENV_PATH, encoding='utf-8') as f:
    for line in f:
        if '=' in line:
            k, v = line.strip().split('=', 1)
            env[k] = v

SUPABASE_URL = env['NEXT_PUBLIC_SUPABASE_URL']
SERVICE_KEY = env['SUPABASE_SERVICE_ROLE_KEY']
HEADERS = {
    'apikey': SERVICE_KEY, 'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json', 'Prefer': 'return=representation',
}

def to_text(v):
    if v is None:
        return None
    return str(v).strip() or None

DATE_RE = re.compile(r'(\d{4})[.\-](\d{2})[.\-](\d{2})')

def to_date_flexible(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()[:10]
    if isinstance(v, str):
        m = DATE_RE.match(v.strip())
        if m:
            return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None

def fetch_json(url):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())

def post_json(url, payload):
    req = urllib.request.Request(url, data=json.dumps(payload).encode('utf-8'), method='POST', headers=HEADERS)
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())

def fetch_all(url_base):
    """페이지네이션 없이 전체를 가져오기 위한 헬퍼 (limit/offset 반복)"""
    results, offset = [], 0
    while True:
        data = fetch_json(f"{url_base}&limit=1000&offset={offset}")
        if not data:
            break
        results.extend(data)
        offset += 1000
    return results

def build_alias_map(wb, franchisee_sheet_name):
    ws = wb[franchisee_sheet_name]
    m = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None or not row[7]:
            continue
        m[str(row[3]).strip()] = str(row[7]).strip()
    return m

def migrate_corp_drivers(driver_sheet_name, franchisee_sheet_name, region, converted_idx=16, note_idx=17):
    """
    서울법인 시트는 성별 컬럼이 없어 전환일=16, 비고=17이지만, 경기법인 시트는 성별 컬럼이
    14번에 끼어 있어 전환일=17, 비고=18로 한 칸씩 밀린다.
    """
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    alias_map = build_alias_map(wb, franchisee_sheet_name)

    r = urllib.parse.quote(region)
    franchisees = fetch_all(f"{SUPABASE_URL}/rest/v1/franchisees?region=eq.{r}&type=eq.%EB%B2%95%EC%9D%B8&select=id,business_name")
    fid_map = {f['business_name']: f['id'] for f in franchisees}

    vehicles = fetch_all(f"{SUPABASE_URL}/rest/v1/vehicles?select=id,plate_no,franchisee_id")
    vid_map = {v['plate_no']: v['id'] for v in vehicles}

    ws = wb[driver_sheet_name]
    batch = []
    inserted, skipped = 0, 0
    log_path = os.path.join(os.path.dirname(__file__), 'migrate_log.txt')
    with open(log_path, 'a', encoding='utf-8') as log:
        log.write(f"\n=== {driver_sheet_name} ===\n")

        def flush():
            nonlocal inserted, skipped
            if not batch:
                return
            try:
                post_json(f"{SUPABASE_URL}/rest/v1/drivers", batch)
                inserted += len(batch)
            except Exception as e:
                skipped += len(batch)
                log.write(f"배치 삽입 실패({len(batch)}건): {e}\n")
            batch.clear()

        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[1] is None or not row[6]:
                skipped += 1
                continue
            alias = str(row[5]).strip() if row[5] else None
            business_name = alias_map.get(alias) if alias else None
            franchisee_id = fid_map.get(business_name) if business_name else None
            if not franchisee_id:
                skipped += 1
                log.write(f"가맹점 매칭 실패: {alias} (기사 {row[6]})\n")
                continue

            note = to_text(row[note_idx]) if len(row) > note_idx else None
            terminated = bool(note and '해지' in note)
            plate_no = to_text(row[12])
            vehicle_id = vid_map.get(plate_no) if plate_no else None

            batch.append({
                'franchisee_id': franchisee_id,
                'vehicle_id': vehicle_id,
                'name': to_text(row[6]),
                'phone': to_text(row[10]),
                'taxi_license_no': to_text(row[11]),
                'education_completed_at': to_date_flexible(row[7]),
                'call_status': '해지' if terminated else '전환완료',
                'call_converted_at': to_date_flexible(row[converted_idx]) if len(row) > converted_idx else None,
                'call_termination_reason': note if terminated else None,
            })
            if len(batch) >= 500:
                flush()
        flush()

    print(f'{driver_sheet_name} 완료. 삽입 {inserted}건, 스킵 {skipped}건.')

if __name__ == '__main__':
    migrate_corp_drivers('가맹콜전환신청(경기법인)', '가맹점 정보(경기법인)', '경기', converted_idx=17, note_idx=18)
