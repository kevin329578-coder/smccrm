import sys, os, json
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import urllib.request
import urllib.parse

XLSX_PATH = r"C:\Users\SMC\Desktop\케빈의 프로젝트\SMC_CRM\캡시 사업운영 관리.xlsx"
ENV_PATH = os.path.join(os.path.dirname(__file__), '..', '.env.local')

env = {}
with open(ENV_PATH, encoding='utf-8') as f:
    for line in f:
        if '=' in line:
            k, v = line.strip().split('=', 1)
            env[k] = v

SUPABASE_URL = env['NEXT_PUBLIC_SUPABASE_URL']
SERVICE_KEY = env['SUPABASE_SERVICE_ROLE_KEY']
HEADERS = {
    'apikey': SERVICE_KEY, 'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type': 'application/json', 'Prefer': 'return=representation',
}

def to_text(v):
    if v is None:
        return None
    return str(v).strip() or None

def fetch_json(url):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())

def post_json(url, payload):
    req = urllib.request.Request(url, data=json.dumps(payload).encode('utf-8'), method='POST', headers=HEADERS)
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())

def build_alias_map(wb, franchisee_sheet_name):
    """가맹점 정보 시트의 '가맹점'(짧은 이름, col 3) -> '사업자명(정식)'(col 7) 매핑"""
    ws = wb[franchisee_sheet_name]
    alias_to_business = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] is None or not row[7]:
            continue
        alias_to_business[str(row[3]).strip()] = str(row[7]).strip()
    return alias_to_business

def fetch_franchisee_id_map(region, ftype):
    r = urllib.parse.quote(region)
    t = urllib.parse.quote(ftype)
    data = fetch_json(f"{SUPABASE_URL}/rest/v1/franchisees?region=eq.{r}&type=eq.{t}&select=id,business_name")
    return {row['business_name']: row['id'] for row in data}

def migrate_vehicle_sheet(vehicle_sheet_name, franchisee_sheet_name, region):
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    alias_map = build_alias_map(wb, franchisee_sheet_name)
    id_map = fetch_franchisee_id_map(region, '법인')

    ws = wb[vehicle_sheet_name]
    batch = []
    inserted, skipped = 0, 0
    log_path = os.path.join(os.path.dirname(__file__), 'migrate_log.txt')
    with open(log_path, 'a', encoding='utf-8') as log:
        log.write(f"\n=== {vehicle_sheet_name} ===\n")

        def flush():
            nonlocal inserted, skipped
            if not batch:
                return
            try:
                post_json(f"{SUPABASE_URL}/rest/v1/vehicles", batch)
                inserted += len(batch)
            except Exception as e:
                skipped += len(batch)
                log.write(f"배치 삽입 실패({len(batch)}건): {e}\n")
            batch.clear()

        for row in ws.iter_rows(min_row=6, values_only=True):
            if row[0] is None or not row[3] or not row[4]:
                skipped += 1
                continue
            alias = str(row[3]).strip()
            business_name = alias_map.get(alias)
            franchisee_id = id_map.get(business_name) if business_name else None
            if not franchisee_id:
                skipped += 1
                log.write(f"가맹점 매칭 실패: {alias} (차량 {row[4]})\n")
                continue
            batch.append({
                'franchisee_id': franchisee_id,
                'plate_no': to_text(row[4]),
                'car_model': to_text(row[8]),
                'decal_type': to_text(row[9]),
                'meter_point': to_text(row[11]),
                'is_new_or_converted': to_text(row[6]),
                'added_at': to_text(row[13]),
                'status': '운행중',
            })
            if len(batch) >= 500:
                flush()
        flush()

    print(f'{vehicle_sheet_name} 완료. 삽입 {inserted}건, 스킵 {skipped}건.')

if __name__ == '__main__':
    migrate_vehicle_sheet('증차차량내역(경기법인)', '가맹점 정보(경기법인)', '경기')
